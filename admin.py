import telebot
from telebot import types
import openpyxl

# Установите токен вашего бота
bot = telebot.TeleBot('7199324736:AAG_j6wOSURnCmVMd6wewJ2zcs_e2GR5LYs')

# Функция для считывания данных из Excel файла
def read_excel_data(file_path, tariff):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if tariff is None or row[2] == tariff:
            data.append(row)
    return data

# Функция для подсчета количества пользователей в каждом тарифе
def count_users_by_tariff(data):
    counts = {}
    for row in data:
        tariff = row[2]
        counts[tariff] = counts.get(tariff, 0) + 1
    return counts

# Функция для создания инлайн-кнопок основного меню
def create_inline_menu_keyboard():
    keyboard = types.InlineKeyboardMarkup()
    buttons = [
        types.InlineKeyboardButton("Активные пользователи", callback_data="active_users"),
        types.InlineKeyboardButton("Редактировать данные", callback_data="edit_data")
    ]
    keyboard.add(*buttons)
    return keyboard

# Функция для создания сообщения с двумя инлайн кнопками
def edit_data_message():
    keyboard = types.InlineKeyboardMarkup()
    buttons = [
        types.InlineKeyboardButton("Добавить ключ", callback_data="edit_add_key"),
        types.InlineKeyboardButton("Добавить пользователя", callback_data="edit_button2")
    ]
    keyboard.add(*buttons)
    return "Выберите действие для редактирования данных:", keyboard

# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    welcome_message = (
        "Привет!\n"
        "Это бот для работы с данными."
    )
    bot.send_message(message.chat.id, welcome_message, reply_markup=create_inline_menu_keyboard())

# Обработчик колбэк-запросов
@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    if call.data == 'active_users':
        active_users_keyboard = types.InlineKeyboardMarkup()
        file_path = 'TELEGA.xlsx'  # Путь к вашему Excel файлу
        tariffs = ["Тариф 1 месяц", "Тариф 3 месяца", "Тариф 6 месяцев", "Свои", "Пустые ключи"]
        users_count = count_users_by_tariff(read_excel_data(file_path, None))  # Вызываем функцию count_users_by_tariff
        message = "Количество пользователей по тарифам:\n"
        for tariff, count in users_count.items():
            if tariff == "Пустые ключи":
                button_text = "Пустые ключи"
            else:
                button_text = tariff
            message += f"{button_text}: {count} пользователей\n"
            if tariff != "Пустые ключи":
                if tariff is not None:  # Добавляем проверку на None
                    active_users_keyboard.add(
                        types.InlineKeyboardButton(button_text, callback_data=f"tariff_{tariff.replace(' ', '_')}"))
        active_users_keyboard.add(types.InlineKeyboardButton("Пустые ключи", callback_data="tariff_empty"))
        bot.send_message(call.message.chat.id, message, reply_markup=active_users_keyboard)

    elif call.data == 'edit_data':
        message, keyboard = edit_data_message()
        bot.send_message(call.message.chat.id, message, reply_markup=keyboard)

    elif call.data == 'edit_add_key':
        bot.send_message(call.message.chat.id, "Введите ключи:")
        bot.register_next_step_handler(call.message, process_keys_input)

    elif call.data == 'edit_button2':
        bot.send_message(call.message.chat.id, "Введите данные в формате USER ID TARIFF TIME:")
        bot.register_next_step_handler(call.message, process_user_data_input)

    elif call.data == 'tariff_Свои':  # Обработка выбора тарифа "Свои"
        file_path = 'TELEGA.xlsx'  # Путь к вашему Excel файлу
        data = read_excel_data_custom(file_path)
        if data:
            formatted_messages = []
            for row in data:
                user = row[0]
                user_id = row[1]
                tariff = row[2]
                purchase_date = row[3]
                key = row[6]
                formatted_message = (
                    f"USER: @{user}\n"
                    f"ID: {user_id}\n"
                    f"ТАРИФ: {tariff}\n"
                    f"Дата покупки: {purchase_date}\n"
                    f"Ключ: {key}\n"
                )
                formatted_messages.append(formatted_message)
            for message in formatted_messages:
                bot.send_message(call.message.chat.id, message)
        else:
            message = "Нет данных для тарифа 'Свои'"
            bot.send_message(call.message.chat.id, message)

    elif call.data.startswith('tariff_'):
        selected_tariff = call.data.replace('tariff_', '').replace('_', ' ')
        file_path = 'TELEGA.xlsx'  # Путь к вашему Excel файлу
        if selected_tariff == "Пустые ключи":
            data = read_excel_data_empty_key(file_path)
        else:
            data = read_excel_data(file_path, selected_tariff)
        if data:
            formatted_messages = []
            for row in data:
                user = row[0]
                user_id = row[1]
                tariff = row[2]
                purchase_date = row[3]
                key = row[6]
                formatted_message = (
                    f"USER: @{user}\n"
                    f"ID: {user_id}\n"
                    f"ТАРИФ: {tariff}\n"
                    f"Дата покупки: {purchase_date}\n"
                    f"Ключ: {key}\n"
                )
                formatted_messages.append(formatted_message)
            for message in formatted_messages:
                bot.send_message(call.message.chat.id, message)
        else:
            message = "Нет данных для выбранного тарифа"
            bot.send_message(call.message.chat.id, message)

# Обработчик для получения введенных данных о пользователе
def process_user_data_input(message):
    user_data = message.text.strip().split()  # Разделяем текст на отдельные слова
    if len(user_data) >= 4:
        user = user_data[0]
        user_id = user_data[1]
        # Объединяем элементы, начиная с третьего, чтобы получить полное название тарифа
        tariff = ' '.join(user_data[2:-1])
        time = user_data[-1]

        file_path = 'TELEGA.xlsx'  # Путь к вашему Excel файлу
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Находим первую пустую строку в колонке A (USER)
        empty_row = 1
        while sheet.cell(row=empty_row, column=1).value is not None:
            empty_row += 1

        # Записываем данные в найденную пустую строку
        sheet.cell(row=empty_row, column=1, value=user)  # USER
        sheet.cell(row=empty_row, column=2, value=user_id)  # ID
        sheet.cell(row=empty_row, column=3, value=tariff)  # TARIFF
        sheet.cell(row=empty_row, column=4, value=time)  # TIME

        wb.save(file_path)
        reply_message = "Данные пользователя успешно добавлены в таблицу!"
        bot.send_message(message.chat.id, reply_message)
    else:
        bot.send_message(message.chat.id, "Некорректный формат данных. Пожалуйста, введите данные в формате USER ID \"TARIFF\" TIME.")

# Обработчик для получения введенных ключей
def process_keys_input(message):
    keys = message.text.strip().split('\n')  # Разделяем текст на отдельные строки
    file_path = 'TELEGA.xlsx'  # Путь к вашему Excel файлу
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    last_row = sheet.max_row  # Находим номер последней строки
    for key in keys:
        sheet.cell(row=last_row + 1, column=7, value=key)  # Добавляем ключ в следующую строку
        last_row += 1
    wb.save(file_path)
    reply_message = "Ключи успешно добавлены в следующие доступные ячейки!"
    bot.send_message(message.chat.id, reply_message)

# Функция для чтения данных только с пустыми ключами
def read_excel_data_empty_key(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[6] and not any(row[:6]):
            data.append(row)
    return data

# Функция для чтения данных только с тарифом "Свои"
def read_excel_data_custom(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == "Свои":
            data.append(row)
    return data

# Запуск бота
bot.polling()