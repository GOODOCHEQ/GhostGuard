import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from openpyxl import load_workbook
from telebot import types
import openpyxl
from datetime import datetime, timedelta


# Установите токен вашего бота
bot = telebot.TeleBot('5797908031:AAEZGcttLY2rBbm-3jQxGZic8HJ2rANUQMU')

# Путь к файлу Excel
excel_file_path = 'TELEGA.xlsx'

# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    # Текст приветствия
    welcome_text = """Привет! Вот что может этот бот.

О нас: Узнайте больше о нашем сервисе VPN.

Полезная информация: Получите информацию, которая поможет вам понять, как все работает.

Мои заказы: Просмотрите данные о своих заказах и подписках.

Поддержка: Обратитесь в службу поддержки для получения помощи и ответов на ваши вопросы.
"""
    # Отправляем сообщение с приветствием и клавиатурой
    bot.send_message(message.chat.id, welcome_text, reply_markup=create_menu_keyboard())

# Обработчик команды "Купить"
@bot.message_handler(func=lambda message: message.text == 'Как оплатить')
def buy(message):
    # Создаем инлайн-кнопки с разными тарифами
    inline_keyboard = InlineKeyboardMarkup()
    # Добавляем кнопки с тарифами
    inline_keyboard.row(
        InlineKeyboardButton("Тариф 1 месяц", callback_data="tariff_A"),
        InlineKeyboardButton("Тариф 3 месяца", callback_data="tariff_B")
    )
    inline_keyboard.row(
        InlineKeyboardButton("Тариф 6 месяцев", callback_data="tariff_C"),
        InlineKeyboardButton("Тариф индивидуальный", callback_data="tariff_D")
    )

    # Отправляем фото и кнопки с тарифами
    bot.send_photo(chat_id=message.chat.id, photo=open('photo_2023-10-30_09-13-28.jpg', 'rb'),
                   caption="Выберите тариф:", reply_markup=inline_keyboard)


# Обработчик нажатия на инлайн-кнопки тарифов
@bot.callback_query_handler(func=lambda call: call.data.startswith('tariff_'))
def process_tariff_callback(callback_query):
    tariff_info = {
        "tariff_A": {"title": "Тариф 1 месяц", "description": "Доступ к VPN на 1 месяц", "price": 19900},
        "tariff_B": {"title": "Тариф 3 месяца", "description": "Доступ к VPN на 3 месяца", "price": 49900},
        "tariff_C": {"title": "Тариф 6 месяцев", "description": "Доступ к VPN на 6 месяцев", "price": 89900},
        "tariff_D": {"title": "Тариф индивидуальный", "description": "Супер тариф с особыми возможностями", "price": 999999}
    }
    tariff_id = callback_query.data
    selected_tariff = tariff_info[tariff_id]

    if tariff_id == "tariff_D":
        # Создаем инлайн-клавиатуру с сылкой
        inline_keyboard = [[InlineKeyboardButton("Написать в поддержку", url="https://t.me/ghostguardvpn")]]
        reply_markup = InlineKeyboardMarkup(inline_keyboard)

        # Отправляем текст и инлайн-ссылку вместе
        bot.send_message(callback_query.from_user.id,
                         "Для подбора сервера под ваши нужды напишите в поддержку с описанием вашей задачи",
                         reply_markup=reply_markup)
    else:
        # Отправляем инвойс для оплаты
        bot.send_invoice(callback_query.from_user.id,
                         title=selected_tariff["title"],
                         description=selected_tariff["description"],
                         provider_token='390540012:LIVE:47653',  # Замените на ваш провайдерский токен
                         currency='RUB',  # Валюта (можно изменить на другую поддерживаемую)
                         prices=[telebot.types.LabeledPrice(label='Цена', amount=selected_tariff["price"])],
                         invoice_payload='payload')


# Обработчик успешной оплаты
@bot.message_handler(content_types=['successful_payment'])
def successful_payment(message):
    # Получение информации о платеже
    payment_info = message.successful_payment

    # Получаем данные о пользователе
    user_id = message.chat.id
    username = message.chat.username

    # Сохраняем данные о пользователе и оплате
    save_payment_data(user_id, username, payment_info)

# Функция для сохранения данных о пользователе, тарифе и дате оплаты
def save_payment_data(user_id, username, tariff):
    # Проверяем, существует ли уже файл Excel
    try:
        # Если файл существует, пробуем загрузить его
        wb = load_workbook(excel_file_path)
        sheet = wb.active
    except FileNotFoundError:
        # Если файл не существует, создаем новый и добавляем заголовки
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'Username'
        sheet['B1'] = 'User ID'
        sheet['C1'] = 'Тариф'
        sheet['D1'] = 'Дата оплаты'

    # Находим первую свободную строку для записи новых данных
    next_row = 1
    while sheet[f'B{next_row}'].value:  # Пока столбец B в строке next_row не пустой
        next_row += 1

    # Записываем данные о пользователе, тарифе и дате оплаты в файл Excel
    sheet[f'A{next_row}'] = username
    sheet[f'B{next_row}'] = user_id
    # Преобразуем идентификатор тарифа в название тарифа
    tariff_names = {
        "tariff_A": "Тариф 1 месяц",
        "tariff_B": "Тариф 3 месяца",
        "tariff_C": "Тариф 6 месяцев",
        "tariff_D": "Тариф индивидуальный"
    }
    sheet[f'C{next_row}'] = tariff_names.get(tariff, "Unknown Tariff")
    sheet[f'D{next_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Сохраняем изменения в файле Excel
    wb.save(excel_file_path)

    # Создаем инлайн-кнопку "ИНСТРУКЦИЯ" с нужной ссылкой
    inline_keyboard = InlineKeyboardMarkup()
    inline_keyboard.add(InlineKeyboardButton("Как подключиться", url="https://teletype.in/@ghostguardvpn/xv_6scteadf"))

    # Отправляем сообщение с текстом и кнопкой "ИНСТРУКЦИЯ"
    bot.send_message(user_id,
                     "Спасибо за покупку! \nЧтобы увидеть свой токен перейдите в Мои заказы\nДля настройки подключения к вашему серверу по токену нажмите на кнопку ниже",
                     reply_markup=inline_keyboard)



# Обработчик команды "Полезная информация"
@bot.message_handler(func=lambda message: message.text == 'Как это работает')
def send_useful_info(message):
    # Создаем клавиатуру с встроенными ссылками
    inline_keyboard = InlineKeyboardMarkup()
    # Добавляем кнопки с встроенными ссылками
    inline_keyboard.row(InlineKeyboardButton("Частые вопросы", url="https://teletype.in/@ghostguardvpn/FxpP9tTdJB8"))
    inline_keyboard.row(
        InlineKeyboardButton("Подключение через OUTLINE", url="https://teletype.in/@ghostguardvpn/xv_6scteadf"))
    inline_keyboard.row(
        InlineKeyboardButton("Политика конфиденциальности", url="https://teletype.in/@ghostguardvpn/f1U9AUIksea"))

    # Отправляем сообщение с клавиатурой
    bot.send_message(message.chat.id, """🔑 Как это работает?

Просто выберите нужный вам план, приобретите ключ через нашего бота и наслаждайтесь безопасным и быстрым интернетом, настроенным под ваши потребности.

Выбирайте наш VPN-сервис для индивидуального подхода к вашей безопасности и скорости в сети!

🚀 Почему выбирают нас?

Индивидуальный подход: Мы предлагаем настройку сервера в соответствии с вашими требованиями и предпочтениями, обеспечивая уникальное решение для каждого пользователя.

Безопасность прежде всего: Наш VPN-сервис обеспечивает защиту вашей конфиденциальности и безопасность вашей передачи данных в сети.

Высокая скорость: Наслаждайтесь высокой скоростью интернета без ограничений и задержек.

Простота и удобство: Приобретайте ключи и настраивайте сервер прямо через нашего бота, с минимальными усилиями и максимальным удовольствием.


Если у вас возникли сложности с подключением - напишите в Поддержку.

Более подробно о работе нашего сервиса и настройках можете узнать из команд ниже:""", reply_markup=inline_keyboard)


# Обработчик команды "Мои заказы"
@bot.message_handler(func=lambda message: message.text == 'Мой VPN')
def my_orders(message):
    # Получаем id пользователя
    user_id = message.from_user.id

    # Пути к файлам Excel
    excel_file_paths = ['TELEGA.xlsx', 'users.xlsx']

    # Создаем список для хранения всех заказов пользователя
    user_orders = []

    # Итерируем по каждому файлу Excel
    for excel_file_path in excel_file_paths:
        try:
            # Загружаем файл Excel
            wb = openpyxl.load_workbook(excel_file_path)
            sheet = wb.active

            # Ищем все заказы пользователя
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=7):
                if str(row[1].value) == str(user_id):
                    user_orders.append({
                        "Тариф": row[2].value,
                        "Дата оплаты": row[3].value,  # Не вызываем strftime() здесь
                        "Токен": row[6].value  # Получаем токен из столбца G
                    })
        except FileNotFoundError:
            # Если файл не найден, продолжаем итерацию
            continue

    # Если найдены заказы пользователя
    if user_orders:
        # Получаем информацию о последнем заказе пользователя
        latest_order = user_orders[-1]
        tariff = latest_order['Тариф']
        payment_date_str = latest_order['Дата оплаты']
        token = latest_order['Токен']

        # Преобразуем строку даты оплаты в объект datetime
        payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d %H:%M:%S')

        # Вычисляем дату окончания срока действия тарифа
        if tariff == "Тариф 1 месяц":
            end_date = payment_date + timedelta(days=30)
        elif tariff == "Тариф 3 месяца":
            end_date = payment_date + timedelta(days=90)
        elif tariff == "Тариф 6 месяцев":
            end_date = payment_date + timedelta(days=180)
        elif tariff == "Тест":
            end_date = payment_date + timedelta(days=5)
        else:
            end_date = payment_date  # Если тариф неизвестен, просто выводим дату оплаты

        # Формируем текст сообщения о заказе и дате окончания
        order_text = f"Ваш тариф: {tariff}\nДействует до: {end_date.strftime('%Y-%m-%d %H:%M:%S')}\n Ваш токен:"

        # Отправляем сообщение с информацией о заказе и дате окончания
        bot.send_message(user_id, order_text)

        # Отправляем вторым сообщением тестовый токен
        bot.send_message(user_id, f"{token}")
    else:
        bot.send_message(user_id, "У вас нет заказов.")


# Обработчик команды "Поддержка"
@bot.message_handler(func=lambda message: message.text == 'Поддержка')
def support(message):
    # Создаем инлайн-кнопку для перехода к диалогу с пользователем @gorponk
    inline_keyboard = InlineKeyboardMarkup()
    inline_keyboard.add(InlineKeyboardButton("Написать в поддержку", url="https://t.me/ghostguardvpn"))

    # Отправляем сообщение с текстом и кнопкой "Написать в поддержку"
    bot.send_message(message.chat.id,
                     "🤝 Наша команда поддержки доступна с 10:00 до 22:00 по МСК  каждый день. Пишите нам по любым вопросам, и мы быстро найдем решение, чтобы обеспечить вам максимальный комфорт и уверенность в использовании нашего сервиса.",
                     reply_markup=inline_keyboard)


# Обработчик команды "О нас"
@bot.message_handler(func=lambda message: message.text == 'О нас')
def about_us(message):
    # Текст о вашей компании
    about_text = """🔒 Наш сервис VPN базируется на технологии OUTLINE, обеспечивая надежное и простое подключение к защищенной сети.

🚀 Скорость и безлимитный трафик – это наш стандарт. Мы используем серверы с каналами до 10 Гбит, что гарантирует плавную работу даже при просмотре YouTube в 4К без каких-либо задержек на любых устройствах.

🕶️ Мы ценим вашу анонимность. Никаких данных о вас не собирается – это наш принцип. Мы предоставляем полную анонимность, не храня историю ваших онлайн-переходов. Ваши данные шифруются и не доступны для просмотра даже администраторам сервиса.

🔐 Кроме того, наш VPN-сервис гарантирует безопасность вашей связи с использованием шифрования трафика и защиты от утечек.

📱 Мы поддерживаем все ваши устройства, включая телефоны и компьютеры. Наша команда всегда готова помочь в любых вопросах, связанных с настройкой или использованием VPN. 

📟Для удобства доступа вы можете подключить любое устройство на любой операционной системе, а так же наш VPN совместим со всеми популярными приложениями"""
    # Отправляем текст о компании
    bot.send_message(chat_id=message.chat.id, text=about_text)

# ПОПРОБОВАТЬ БЕСПЛАТНО

# Функция сохранения данных пользователя в файл Excel
def save_to_excel(username, user_id, action):
    try:
        wb = openpyxl.load_workbook('users.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Username", "User ID", "Action", "Date"])

    # Ищем первую пустую строку в столбце A
    next_row = 1
    while sheet.cell(row=next_row, column=1).value is not None:
        next_row += 1

    # Записываем данные пользователя в найденную строку
    sheet.cell(row=next_row, column=1).value = username
    sheet.cell(row=next_row, column=2).value = user_id
    sheet.cell(row=next_row, column=3).value = action
    sheet.cell(row=next_row, column=4).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb.save('users.xlsx')
    print("Данные пользователя успешно сохранены:", username, user_id, action)

# Функция получения тестового токена пользователя из файла Excel
def get_test_token(username):
    try:
        wb = openpyxl.load_workbook('users.xlsx')
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            if row[0] == username:
                return row[6]  # Возвращаем значение из 6 столбца (после 5 столбца)
    except FileNotFoundError:
        pass

    return None

# Обработчик команды "Попробовать бесплатно"
@bot.message_handler(func=lambda message: message.text == 'Попробовать бесплатно')
def try_for_free(message):
    # Проверяем, есть ли уже данные пользователя в файле
    test_token = get_test_token(message.chat.username)

    if test_token:
        # Если данные пользователя уже есть, отправляем сообщение о том, что он уже воспользовался бесплатным периодом
        bot.send_message(message.chat.id, "Извините, вы уже воспользовались бесплатным периодом.")
    else:
        # Получаем дату начала тестового периода и прибавляем 5 дней
        start_date = datetime.now()
        end_date = start_date + timedelta(days=5)

        # Создаем инлайн-кнопку для инструкции
        markup = types.InlineKeyboardMarkup()
        instruction_button = types.InlineKeyboardButton("Инструкция к подключению", url="https://teletype.in/@ghostguardvpn/xv_6scteadf")
        markup.add(instruction_button)

        # Формируем сообщение о доступе и дате окончания
        message_text = "Теперь у вас есть тестовый доступ!\n\nДействует до: {}".format(end_date.strftime("%Y-%m-%d"))
        message_text += "\n\nВаш токен находится в сообщении снизу. Для подключения следуйте инструкции:"

        # Отправляем сообщение с текстом, кнопкой и инструкцией
        bot.send_message(message.chat.id, message_text, reply_markup=markup)

        # Сохраняем данные пользователя в файл Excel
        save_to_excel(message.chat.username, message.chat.id, "Тест")

        # Отправляем вторым сообщением тестовый токен
        test_token_message = "{}".format(get_test_token(message.chat.username))
        bot.send_message(message.chat.id, test_token_message)


# Функция для создания клавиатуры с меню
def create_menu_keyboard():
    keyboard = ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)  # Установка row_width=2
    buttons = ["О нас", "Как это работает", "Как оплатить", "Попробовать бесплатно", "Мой VPN", "Поддержка"]
    for i in range(0, len(buttons), 2):  # Итерирование по кнопкам попарно
        button_row = buttons[i:i + 2]  # Получение двух кнопок для каждого ряда
        keyboard.add(*[KeyboardButton(text=button) for button in button_row])
    return keyboard

# Запуск бота
bot.polling()
